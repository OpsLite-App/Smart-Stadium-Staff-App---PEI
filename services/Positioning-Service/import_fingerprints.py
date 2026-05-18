import argparse
import csv
import json
from collections import defaultdict
from pathlib import Path

import httpx


def normalize_bssid(bssid: str) -> str:
    return bssid.strip().lower()


def read_xlsx_file(path: Path):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "openpyxl não está instalado. Ative o virtualenv e instale openpyxl."
        ) from exc

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []

    headers = [str(cell).strip().lower() if cell is not None else "" for cell in rows[0]]
    return headers, rows[1:]


def parse_input_file(path: Path):
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        headers, data_rows = read_xlsx_file(path)
        reader = data_rows
        has_header = bool(headers and any(headers))
    else:
        with path.open(newline="", encoding="utf-8") as f:
            sample = f.read(1024)
            f.seek(0)
            sniffer = csv.Sniffer()
            has_header = sniffer.has_header(sample)
            dialect = sniffer.sniff(sample)
            reader = csv.reader(f, dialect)

            if has_header:
                headers = [h.strip().lower() for h in next(reader)]
            else:
                headers = ["location_id", "zone", "x", "y", "bssid", "rssi"]

    rows = []
    for row in reader:
        if not row or all((str(cell).strip() == "") for cell in row):
            continue
        if len(row) < 6:
            raise ValueError(f"Linha inválida, esperados 6 campos: {row}")

        data = dict(zip(headers, row))
        rows.append({
            "location_id": str(data.get("location_id", "")).strip(),
            "zone": str(data.get("zone", "")).strip(),
            "x": float(data.get("x", 0)),
            "y": float(data.get("y", 0)),
            "bssid": normalize_bssid(data.get("bssid", "")),
            "rssi": int(float(data.get("rssi", 0))),
        })

    if not rows:
        raise ValueError("Arquivo de entrada vazio ou inválido.")

    return rows


def build_fingerprints(rows):
    groups = defaultdict(lambda: {"zone": None, "x": None, "y": None, "rssi_map": {}})

    for row in rows:
        key = (row["location_id"], row["zone"], row["x"], row["y"])
        if groups[key]["zone"] is None:
            groups[key]["zone"] = row["zone"]
            groups[key]["x"] = row["x"]
            groups[key]["y"] = row["y"]

        groups[key]["rssi_map"][row["bssid"]] = row["rssi"]

    fingerprints = []
    for (location_id, _, _, _), data in groups.items():
        fingerprints.append({
            "location_id": location_id,
            "zone": data["zone"],
            "x": data["x"],
            "y": data["y"],
            "rssi_map": data["rssi_map"],
        })

    return fingerprints


def post_fingerprint(url: str, fingerprint: dict, client: httpx.Client):
    response = client.post(url, json=fingerprint, timeout=30.0)
    response.raise_for_status()
    return response.json()


def main():
    parser = argparse.ArgumentParser(
        description="Importa um dataset de fingerprints de APs para o Positioning-Service."
    )
    parser.add_argument("input_file", help="Caminho para o ficheiro CSV/TSV/XLSX de access points")
    parser.add_argument(
        "--url",
        default="http://localhost:8000/fingerprints",
        help="Endpoint HTTP do Positioning-Service /fingerprints",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Não envia ao serviço; mostra o número de fingerprints e amostra de JSON.",
    )
    parser.add_argument(
        "--ignore-invalid",
        action="store_true",
        help="Ignora linhas inválidas em vez de falhar imediatamente.",
    )
    args = parser.parse_args()

    path = Path(args.input_file)
    if not path.exists():
        raise FileNotFoundError(f"Ficheiro não encontrado: {path}")

    rows = parse_input_file(path)
    fingerprints = build_fingerprints(rows)

    print(f"Encontrados {len(rows)} leituras, {len(fingerprints)} fingerprints únicos.")
    print("Exemplo de fingerprint:")
    print(json.dumps(fingerprints[0], indent=2, ensure_ascii=False))

    if args.dry_run:
        return

    with httpx.Client() as client:
        for index, fingerprint in enumerate(fingerprints, start=1):
            try:
                result = post_fingerprint(args.url, fingerprint, client)
                print(f"[{index}/{len(fingerprints)}] importado: {result['location_id']}")
            except Exception as exc:
                print(f"Erro ao importar fingerprint {fingerprint['location_id']}: {exc}")
                raise


if __name__ == "__main__":
    main()
